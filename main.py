from openpyxl import load_workbook
import kznumber as kzn
import doclister
import os

# getting KZ number from user
kz_number = kzn.kzNumber()

# setting a path for listing documents
project_path = 'E:/Zajecia/Projekty/14-Inteligentny_pojazd/01-Mechanika'
'C:/Repozytorium/Approved/031-Pxx0/W031-Pxx0'
# 'E:/Zajecia/Projekty/14-Inteligentny_pojazd/01-Mechanika'
# 'E:/Zajecia/Projekty/17-Dir_docs_lister/Dir_docs_lister'

# destination filename
list_name = 'W' + kz_number.kz_I + '-00-Spis_dokumentacji_wyrobu.xlsx'

list_path = project_path + '/' + list_name

# number of row where the list is beginning
begin = 2

# check if a list already exist
if os.path.isfile(list_path):
    # if exist - update a file using KZ number
    print("Update a file")

    wb = load_workbook(filename=list_path)
    sheet = wb['Lista_dokumentów']

    # making existing list of documents for comparison
    existing_list = doclister.makelist([], sheet, begin)

    # making list of new documents for comparison
    new_list = doclister.convtonamelist(doclister.makedoclist(project_path, []))

    doclister.comparelists(existing_list, new_list)

    #for item in new_list:
     #   print(item)


    # updating list of documents
    doclister.exportlist(doclister.makedoclist(project_path, _doclist=[]), sheet, kz_number.kz, begin)

    wb.save(list_path)
else:
    # if not - make one using KZ number
    print("Make a file")

    wb = load_workbook(filename='XXX-Szablon_listy.xlsx')
    sheet = wb['Lista_dokumentów']

    # getting list of documents
    doclister.exportlist(doclister.makedoclist(project_path, _doclist=[]), sheet, kz_number.kz, begin)

    wb.save(list_path)


